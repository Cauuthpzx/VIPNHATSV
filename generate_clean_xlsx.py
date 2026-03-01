"""
Generate a clean API documentation XLSX from the captured data.
Focus ONLY on actual API endpoints, not static resources.
"""
import json
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')

# All API endpoints discovered from the original site
# Each entry has: page, endpoint, method, request params, response structure
API_DOCS = [
    # === WELCOME PAGE ===
    {
        "page": "Welcome (Trang chu)",
        "endpoint": "/agent/welcome",
        "method": "GET",
        "content_type": "text/html",
        "description": "Trang chu - hien thi thong tin dang nhap va so du",
        "request_params": "None (page load)",
        "response_format": "HTML page with login info and balance data embedded",
        "sample_response": "HTML with: lastLoginTime, lastLoginIp, currentLoginTime, currentLoginIp, agentBalance, frozenAmount",
    },

    # === USER LIST ===
    {
        "page": "User List (Danh sach hoi vien)",
        "endpoint": "/agent/user.html",
        "method": "POST",
        "content_type": "application/x-www-form-urlencoded / application/json",
        "description": "Lay danh sach hoi vien (thanh vien)",
        "request_params": json.dumps({
            "page": 1,
            "limit": 10,
            "username": "",
            "user_type": "",
            "status": "",
        }, ensure_ascii=False),
        "response_format": "JSON: {data: [...], count: N, code: 0}",
        "sample_response": json.dumps({
            "data": [{
                "id": 2470509,
                "username": "conghao12",
                "type_format": "Hoi vien chinh thuc",
                "parent_user": "xiu777",
                "balance": "0.0000",
                "frozen_amount": "0.0000",
                "status_format": "Binh thuong",
                "last_login_time": "2026-03-01 01:33:06",
                "last_login_ip": "x.x.x.x",
                "created_at": "2025-12-20 10:00:00"
            }],
            "count": 100,
            "code": 0
        }, indent=2, ensure_ascii=False),
    },

    # === INVITE LIST ===
    {
        "page": "Invite List (Ma gioi thieu)",
        "endpoint": "/agent/inviteList.html",
        "method": "POST",
        "content_type": "application/json",
        "description": "Lay danh sach ma gioi thieu",
        "request_params": json.dumps({
            "page": 1,
            "limit": 10,
            "create_time": "",
            "user_register_time": "",
            "invite_code": ""
        }, ensure_ascii=False),
        "response_format": "JSON: {data: [...], count: N, code: 0}",
        "sample_response": json.dumps({
            "data": [{
                "id": 102246,
                "invite_code": "6926432",
                "user_type": "Hoi vien thuong",
                "reg_count": 20150,
                "scope_reg_count": 20150,
                "recharge_count": 10022,
                "first_recharge_count": 10022,
                "register_recharge_count": 8067,
                "url": "https://...",
                "created_at": "2025-01-01"
            }],
            "count": 5,
            "code": 0
        }, indent=2, ensure_ascii=False),
    },
    {
        "page": "Invite List (Ma gioi thieu)",
        "endpoint": "/agent/inviteList.html",
        "method": "POST (form-urlencoded)",
        "content_type": "application/x-www-form-urlencoded",
        "description": "Lay danh sach ma gioi thieu (form submit)",
        "request_params": "page=1&limit=10&create_time=&user_register_time=&invite_code=",
        "response_format": "JSON: {data: [...], count: N, code: 0}",
        "sample_response": "(same as JSON version above)",
    },

    # === REPORT LOTTERY ===
    {
        "page": "Report Lottery (Bao cao xo so)",
        "endpoint": "/agent/reportLottery.html",
        "method": "POST",
        "content_type": "application/json",
        "description": "Bao cao xo so - thong ke cuoc xo so theo ngay",
        "request_params": json.dumps({
            "page": 1,
            "limit": 10,
            "date": "2026-03-01 | 2026-03-01",
            "lottery_id": "",
            "username": ""
        }, ensure_ascii=False),
        "response_format": "JSON: {data: [...], count: N, code: 0, total_data: {...}}",
        "sample_response": json.dumps({
            "data": [{
                "username": "tuanbeo1202",
                "user_parent_format": "xiu777",
                "bet_count": "38",
                "bet_amount": "6727000.0000",
                "valid_amount": "6727000.0000",
                "rebate_amount": "0.0000",
                "result": "-219700.0000",
                "win_lose": "-219700.0000",
                "win_amount": "6507300.0000",
                "lottery_name": "Ten loai xo"
            }],
            "count": 50,
            "code": 0,
            "total_data": {
                "total_bet_count": "763",
                "total_bet_amount": "131309415.0000",
                "total_valid_amount": "131309415.0000",
                "total_rebate_amount": "0.0000",
                "total_result": "-10481127.2220",
                "total_win_lose": "-10481127.2220",
                "total_win_amount": "120828287.7780",
                "total_players": "25"
            }
        }, indent=2, ensure_ascii=False),
    },

    # === REPORT FUNDS ===
    {
        "page": "Report Funds (Sao ke giao dich)",
        "endpoint": "/agent/reportFunds.html",
        "method": "POST",
        "content_type": "application/json",
        "description": "Sao ke giao dich - bao cao nap/rut theo ngay",
        "request_params": json.dumps({
            "page": 1,
            "limit": 10,
            "date": "2026-03-01 | 2026-03-01",
            "username": ""
        }, ensure_ascii=False),
        "response_format": "JSON: {data: [...], count: N, code: 0, total_data: {...}}",
        "sample_response": json.dumps({
            "data": [{
                "username": "hoangnhan79",
                "user_parent_format": "xiu777",
                "deposit_count": 1,
                "deposit_amount": "3500000.0000",
                "withdrawal_count": 0,
                "withdrawal_amount": "0.0000",
                "charge_fee": "0.0000",
                "agent_commission": "0.0000",
                "net_amount": "3500000.0000"
            }],
            "count": 10,
            "code": 0,
            "total_data": {
                "total_deposit_amount": "50000000.0000",
                "total_withdrawal_amount": "30000000.0000",
                "total_charge_fee": "0.0000"
            }
        }, indent=2, ensure_ascii=False),
    },

    # === REPORT THIRD GAME ===
    {
        "page": "Report Third Game (Bao cao nha cung cap)",
        "endpoint": "/agent/reportThirdGame.html",
        "method": "POST",
        "content_type": "application/json",
        "description": "Bao cao nha cung cap (game ben thu 3)",
        "request_params": json.dumps({
            "page": 1,
            "limit": 10,
            "date": "2026-03-01 | 2026-03-01",
            "username": "",
            "platform_id": ""
        }, ensure_ascii=False),
        "response_format": "JSON: {data: [...], count: N, code: 0, total_data: {...}}",
        "sample_response": json.dumps({
            "data": [{
                "username": "phuochien2685",
                "platform_id_name": "AE",
                "t_bet_times": "1",
                "t_bet_amount": "35000.0000",
                "t_turnover": "35000.0000",
                "t_prize": "0.0000",
                "t_win_lose": "-35000.0000"
            }],
            "count": 20,
            "code": 0,
            "total_data": {
                "total_bet_times": "100",
                "total_bet_amount": "5000000.0000",
                "total_turnover": "5000000.0000",
                "total_prize": "2000000.0000",
                "total_win_lose": "-3000000.0000"
            }
        }, indent=2, ensure_ascii=False),
    },

    # === BANK LIST ===
    {
        "page": "Bank List (The ngan hang)",
        "endpoint": "/agent/bankList.html",
        "method": "POST",
        "content_type": "application/x-www-form-urlencoded",
        "description": "Danh sach the ngan hang cua hoi vien",
        "request_params": json.dumps({
            "page": 1,
            "limit": 10,
            "username": "",
            "bank_name": "",
            "card_no": ""
        }, ensure_ascii=False),
        "response_format": "JSON: {data: [...], count: N, code: 0}",
        "sample_response": json.dumps({
            "data": [{
                "id": 1,
                "username": "user1",
                "bank_name": "Vietcombank",
                "card_no": "1234****5678",
                "real_name": "NGUYEN VAN A",
                "created_at": "2025-12-01"
            }],
            "count": 5,
            "code": 0
        }, indent=2, ensure_ascii=False),
    },

    # === DEPOSIT LIST ===
    {
        "page": "Deposit List (Nap tien)",
        "endpoint": "/agent/depositAndWithdrawal.html",
        "method": "POST",
        "content_type": "application/x-www-form-urlencoded",
        "description": "Danh sach nap tien",
        "request_params": json.dumps({
            "page": 1,
            "limit": 10,
            "username": "",
            "type": "",
            "status": "",
            "date": "2026-03-01 | 2026-03-01"
        }, ensure_ascii=False),
        "response_format": "JSON: {data: [...], count: N, code: 0}",
        "sample_response": json.dumps({
            "data": [{
                "username": "phuochien2685",
                "user_parent_format": "xiu777",
                "amount": "500000.0000",
                "type": "Nap tien",
                "status": "Thanh cong",
                "created_at": "2026-03-01 01:00:00"
            }],
            "count": 50,
            "code": 0
        }, indent=2, ensure_ascii=False),
    },

    # === WITHDRAWALS RECORD ===
    {
        "page": "Withdrawals Record (Lich su rut tien)",
        "endpoint": "/agent/withdrawalsRecord.html",
        "method": "POST",
        "content_type": "application/x-www-form-urlencoded",
        "description": "Lich su rut tien",
        "request_params": json.dumps({
            "page": 1,
            "limit": 10,
            "username": "",
            "status": "",
            "date": "2026-03-01 | 2026-03-01"
        }, ensure_ascii=False),
        "response_format": "JSON: {data: [...], count: N, code: 0}",
        "sample_response": json.dumps({
            "data": [{
                "serial_no": "1772308224165710055721",
                "create_time": "2026-03-01 02:50:24",
                "username": "phamh...",
                "amount": "1000000.0000",
                "status": "Cho xu ly",
                "bank_name": "Vietcombank",
                "card_no": "1234****5678"
            }],
            "count": 20,
            "code": 0
        }, indent=2, ensure_ascii=False),
    },

    # === WITHDRAW ===
    {
        "page": "Withdraw (Rut tien)",
        "endpoint": "/agent/withdraw.html",
        "method": "POST",
        "content_type": "application/x-www-form-urlencoded",
        "description": "Thuc hien rut tien (submit form)",
        "request_params": json.dumps({
            "amount": "1000000",
            "fund_password": "******",
            "bank_id": "1"
        }, ensure_ascii=False),
        "response_format": "JSON: {code: 0/1, msg: 'success/error message'}",
        "sample_response": json.dumps({
            "code": 0,
            "msg": "Rut tien thanh cong"
        }, indent=2, ensure_ascii=False),
    },

    # === BET LIST ===
    {
        "page": "Bet List (Don cuoc)",
        "endpoint": "/agent/bet.html",
        "method": "POST",
        "content_type": "application/json",
        "description": "Danh sach don cuoc xo so (phan trang)",
        "request_params": json.dumps({
            "page": 1,
            "limit": 10,
            "create_time": "2026-03-01 | 2026-03-02",
            "date": "2026-03-01 | 2026-03-01",
            "username": "",
            "serial_no": "",
            "lottery_id": "",
            "play_type_id": "",
            "play_id": "",
            "status": "",
            "es": 1
        }, ensure_ascii=False),
        "response_format": "JSON: {data: [...], count: N, code: 0}",
        "sample_response": json.dumps({
            "data": [{
                "serial_no": "17723100981244963072680951158",
                "username": "elizamy6868",
                "create_time": "2026-03-01 03:21:38",
                "lottery_name": "Mien Bac VIP 75 giay",
                "play_type_name": "[So bon]Keo doi",
                "play_name": "Keo doi",
                "content": "01",
                "money": "5000.0000",
                "odds": "11.0000",
                "result": "-5000.0000",
                "status": "Khong trung",
                "rebate_amount": "0.0000"
            }],
            "count": 763,
            "code": 0
        }, indent=2, ensure_ascii=False),
    },
    {
        "page": "Bet List (Don cuoc) - SUMMARY",
        "endpoint": "/agent/bet.html",
        "method": "POST",
        "content_type": "application/json",
        "description": "Thong ke tong hop don cuoc (is_summary=1)",
        "request_params": json.dumps({
            "create_time": "2026-03-01 | 2026-03-02",
            "date": "2026-03-01 | 2026-03-01",
            "username": "",
            "serial_no": "",
            "lottery_id": "",
            "play_type_id": "",
            "play_id": "",
            "status": "",
            "is_summary": 1,
            "es": 1
        }, ensure_ascii=False),
        "response_format": "JSON: {data: [], count: N, code: 0, total_data: {...}}",
        "sample_response": json.dumps({
            "data": [],
            "count": 763,
            "code": 0,
            "create_time": "2026-03-01 | 2026-03-02",
            "total_data": {
                "total_result": "-10,481,127.2220",
                "total_money": "131,309,415.0000",
                "total_rebate_amount": "0.0000",
                "total_count": 763
            }
        }, indent=2, ensure_ascii=False),
    },

    # === BET ORDER (3rd party) ===
    {
        "page": "Bet Order (Don cuoc thu 3)",
        "endpoint": "/agent/betOrder.html",
        "method": "POST",
        "content_type": "application/json",
        "description": "Don cuoc ben thu 3 (game bai, no hu, etc.)",
        "request_params": json.dumps({
            "page": 1,
            "limit": 10,
            "bet_time": "2026-03-01 | 2026-03-02",
            "serial_no": "",
            "platform_username": "",
            "es": 1
        }, ensure_ascii=False),
        "response_format": "JSON: {data: [...], count: N, code: 0}",
        "sample_response": json.dumps({
            "data": [{
                "serial_no": "1942309934086632036",
                "platform_id_name": "PG NEW",
                "platform_username": "kisski",
                "c_name": "No hu",
                "game_name": "Wild Bounty Showdown",
                "bet_amount": 90000,
                "turnover": 90000,
                "prize": 0,
                "win_lose": -90000,
                "bet_time": "2026-03-01 03:00:00"
            }],
            "count": 100,
            "code": 0
        }, indent=2, ensure_ascii=False),
    },

    # === EDIT PASSWORD ===
    {
        "page": "Edit Password (Doi mat khau)",
        "endpoint": "/agent/editPassword.html",
        "method": "POST",
        "content_type": "application/x-www-form-urlencoded",
        "description": "Doi mat khau dang nhap",
        "request_params": json.dumps({
            "old_password": "******",
            "new_password": "******",
            "confirm_password": "******"
        }, ensure_ascii=False),
        "response_format": "JSON: {code: 0/1, msg: 'success/error'}",
        "sample_response": json.dumps({
            "code": 0,
            "msg": "Doi mat khau thanh cong"
        }, indent=2, ensure_ascii=False),
    },

    # === EDIT FUND PASSWORD ===
    {
        "page": "Edit Fund Password (MK giao dich)",
        "endpoint": "/agent/editFundPassword.html",
        "method": "POST",
        "content_type": "application/x-www-form-urlencoded",
        "description": "Doi mat khau giao dich",
        "request_params": json.dumps({
            "old_fund_password": "******",
            "new_fund_password": "******",
            "confirm_fund_password": "******"
        }, ensure_ascii=False),
        "response_format": "JSON: {code: 0/1, msg: 'success/error'}",
        "sample_response": json.dumps({
            "code": 0,
            "msg": "Doi mat khau giao dich thanh cong"
        }, indent=2, ensure_ascii=False),
    },

    # === GET LOTTERY (dropdown data) ===
    {
        "page": "Rebate Odds / Bet List (Dropdown)",
        "endpoint": "/agent/getLottery",
        "method": "POST",
        "content_type": "application/json",
        "description": "Lay danh sach loai xo so cho dropdown/select",
        "request_params": json.dumps({
            "type": "getLottery",
            "series_id": ""
        }, ensure_ascii=False),
        "response_format": "JSON: {code: 1, msg: '', data: {seriesData: [...], lotteryData: [...]}}",
        "sample_response": json.dumps({
            "code": 1,
            "msg": "",
            "data": {
                "seriesData": [
                    {"id": 1, "name": "Mien Nam"},
                    {"id": 2, "name": "Mien Bac"},
                    {"id": 3, "name": "Mien Trung"},
                    {"id": 6, "name": "Xo so nhanh"},
                    {"id": 7, "name": "Keno"},
                    {"id": 8, "name": "Xo so cao"},
                    {"id": 9, "name": "Sicbo"},
                    {"id": 10, "name": "pk"},
                    {"id": 11, "name": "Wingo"}
                ],
                "lotteryData": [
                    {"id": 1, "name": "Lottery Name", "series_id": 1}
                ]
            }
        }, indent=2, ensure_ascii=False),
    },

    # === REBATE ODDS PANEL (init) ===
    {
        "page": "Rebate Odds (Ti le hoan tra) - Init",
        "endpoint": "/agent/getRebateOddsPanel.html",
        "method": "POST",
        "content_type": "application/json",
        "description": "Khoi tao panel ti le hoan tra (type=init)",
        "request_params": json.dumps({
            "type": "init"
        }, ensure_ascii=False),
        "response_format": "JSON: {code: 1, msg: '', data: {seriesData: [...], lotteryData: [...]}}",
        "sample_response": json.dumps({
            "code": 1,
            "msg": "",
            "data": {
                "seriesData": [
                    {"id": 1, "name": "Mien Nam"},
                    {"id": 2, "name": "Mien Bac"}
                ],
                "lotteryData": [
                    {"id": 1, "name": "Lottery 1", "series_id": 1}
                ]
            }
        }, indent=2, ensure_ascii=False),
    },
    {
        "page": "Rebate Odds (Ti le hoan tra) - getLottery",
        "endpoint": "/agent/getRebateOddsPanel.html",
        "method": "POST",
        "content_type": "application/json",
        "description": "Lay ti le hoan tra theo loai xo so (type=getLottery)",
        "request_params": json.dumps({
            "type": "getLottery",
            "series_id": 1
        }, ensure_ascii=False),
        "response_format": "JSON: {code: 1, msg: '', data: {tableHead: [...], tableBody: [...]}}",
        "sample_response": json.dumps({
            "code": 1,
            "msg": "",
            "data": {
                "tableHead": ["Loai cuoc", "Ti le A", "Ti le B"],
                "tableBody": [
                    {"play_name": "2D", "rate_a": "99.5", "rate_b": "98.0"}
                ]
            }
        }, indent=2, ensure_ascii=False),
    },
]

def create_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "API Documentation"

    # Styles
    hfont = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    headers = [
        "STT", "Page / Menu", "API Endpoint", "HTTP Method",
        "Content-Type", "Description",
        "Request Parameters", "Response Format",
        "Sample Response"
    ]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = halign
        c.border = border

    widths = [6, 35, 40, 18, 35, 45, 55, 40, 100]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    dfont = Font(name="Consolas", size=9)
    walign = Alignment(vertical="top", wrap_text=True)
    label_font = Font(name="Arial", size=10)

    current_page = ""
    for ri, api in enumerate(API_DOCS, 2):
        row_data = [
            ri - 1,
            api["page"],
            api["endpoint"],
            api["method"],
            api.get("content_type", ""),
            api.get("description", ""),
            api.get("request_params", ""),
            api.get("response_format", ""),
            api.get("sample_response", ""),
        ]

        # Highlight section changes
        page_base = api["page"].split(" -")[0].split(" (")[0]
        is_new_section = page_base != current_page
        current_page = page_base

        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=str(val))
            if ci in (7, 9):  # Params and Response use monospace
                c.font = dfont
            else:
                c.font = label_font
            c.alignment = walign
            c.border = border
            if is_new_section:
                c.fill = section_fill

    # Freeze
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(API_DOCS) + 1}"

    # === Sheet 2: Summary ===
    ws2 = wb.create_sheet("API Summary")
    summary_headers = ["Page", "Endpoint", "Method", "Key Params"]
    for ci, h in enumerate(summary_headers, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = halign
        c.border = border

    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 50

    for ri, api in enumerate(API_DOCS, 2):
        # Extract key param names
        try:
            params = json.loads(api.get("request_params", "{}"))
            key_params = ", ".join(params.keys()) if isinstance(params, dict) else api.get("request_params", "")
        except (json.JSONDecodeError, TypeError):
            key_params = api.get("request_params", "")

        ws2.cell(row=ri, column=1, value=api["page"]).border = border
        ws2.cell(row=ri, column=2, value=api["endpoint"]).border = border
        ws2.cell(row=ri, column=3, value=api["method"]).border = border
        ws2.cell(row=ri, column=4, value=key_params).border = border

    ws2.freeze_panes = "A2"

    # === Sheet 3: Date Parameters ===
    ws3 = wb.create_sheet("Date Parameters")
    date_headers = ["Page", "Date Field Name", "Date Format", "Example Value", "Notes"]
    for ci, h in enumerate(date_headers, 1):
        c = ws3.cell(row=1, column=ci, value=h)
        c.font = hfont
        c.fill = hfill
        c.border = border

    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 30
    ws3.column_dimensions['D'].width = 35
    ws3.column_dimensions['E'].width = 40

    date_entries = [
        ("Report Lottery", "date", "YYYY-MM-DD | YYYY-MM-DD", "2026-03-01 | 2026-03-01", "Date range separated by ' | '"),
        ("Report Funds", "date", "YYYY-MM-DD | YYYY-MM-DD", "2026-03-01 | 2026-03-01", "Date range separated by ' | '"),
        ("Report Third Game", "date", "YYYY-MM-DD | YYYY-MM-DD", "2026-03-01 | 2026-03-01", "Date range separated by ' | '"),
        ("Bet List", "create_time", "YYYY-MM-DD | YYYY-MM-DD", "2026-03-01 | 2026-03-02", "Thoi gian tao don cuoc"),
        ("Bet List", "date", "YYYY-MM-DD | YYYY-MM-DD", "2026-03-01 | 2026-03-01", "Ngay ket qua xo so"),
        ("Bet Order", "bet_time", "YYYY-MM-DD | YYYY-MM-DD", "2026-03-01 | 2026-03-02", "Thoi gian cuoc game thu 3"),
        ("Invite List", "create_time", "YYYY-MM-DD HH:mm:ss range", "", "Thoi gian tao ma gioi thieu (datetime)"),
        ("Invite List", "user_register_time", "YYYY-MM-DD HH:mm:ss range", "", "Thoi gian dang ky nguoi dung"),
        ("Deposit List", "date", "YYYY-MM-DD | YYYY-MM-DD", "2026-03-01 | 2026-03-01", "Ngay nap tien"),
        ("Withdrawals Record", "date", "YYYY-MM-DD | YYYY-MM-DD", "2026-03-01 | 2026-03-01", "Ngay rut tien"),
    ]

    for ri, (page, field, fmt, example, notes) in enumerate(date_entries, 2):
        ws3.cell(row=ri, column=1, value=page).border = border
        ws3.cell(row=ri, column=2, value=field).border = border
        ws3.cell(row=ri, column=3, value=fmt).border = border
        ws3.cell(row=ri, column=4, value=example).border = border
        ws3.cell(row=ri, column=5, value=notes).border = border

    ws3.freeze_panes = "A2"

    # === Sheet 4: Response Fields ===
    ws4 = wb.create_sheet("Response Fields")
    field_headers = ["Page", "Field Name", "Field Type", "Description", "Example Value"]
    for ci, h in enumerate(field_headers, 1):
        c = ws4.cell(row=1, column=ci, value=h)
        c.font = hfont
        c.fill = hfill
        c.border = border

    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 25
    ws4.column_dimensions['C'].width = 15
    ws4.column_dimensions['D'].width = 40
    ws4.column_dimensions['E'].width = 35

    field_entries = [
        # Common
        ("All pages", "code", "int", "Response code (0=success, 1=success for some)", "0"),
        ("All pages", "msg", "string", "Error/success message", ""),
        ("All pages", "data", "array", "Array of data records", "[...]"),
        ("All pages", "count", "int", "Total number of records", "763"),
        # User
        ("User List", "id", "int", "User ID", "2470509"),
        ("User List", "username", "string", "Username", "conghao12"),
        ("User List", "type_format", "string", "User type label", "Hoi vien chinh thuc"),
        ("User List", "parent_user", "string", "Parent agent", "xiu777"),
        ("User List", "balance", "string", "User balance", "0.0000"),
        ("User List", "frozen_amount", "string", "Frozen amount", "0.0000"),
        ("User List", "status_format", "string", "Status label", "Binh thuong"),
        # Invite
        ("Invite List", "invite_code", "string", "Invite code", "6926432"),
        ("Invite List", "user_type", "string", "User type", "Hoi vien thuong"),
        ("Invite List", "reg_count", "int", "Registration count", "20150"),
        ("Invite List", "recharge_count", "int", "Recharge count", "10022"),
        # Report Lottery
        ("Report Lottery", "bet_count", "string", "Number of bets", "38"),
        ("Report Lottery", "bet_amount", "string", "Bet amount", "6727000.0000"),
        ("Report Lottery", "valid_amount", "string", "Valid bet amount", "6727000.0000"),
        ("Report Lottery", "rebate_amount", "string", "Rebate amount", "0.0000"),
        ("Report Lottery", "result", "string", "Win/loss result", "-219700.0000"),
        ("Report Lottery", "win_lose", "string", "Win/loss", "-219700.0000"),
        ("Report Lottery", "win_amount", "string", "Win amount", "6507300.0000"),
        ("Report Lottery", "lottery_name", "string", "Lottery type name", "Mien Bac VIP"),
        # Bet List
        ("Bet List", "serial_no", "string", "Bet serial number", "17723100981244963072680951158"),
        ("Bet List", "lottery_name", "string", "Lottery name", "Mien Bac VIP 75 giay"),
        ("Bet List", "play_type_name", "string", "Play type", "[So bon]Keo doi"),
        ("Bet List", "play_name", "string", "Play name", "Keo doi"),
        ("Bet List", "content", "string", "Bet content/number", "01"),
        ("Bet List", "money", "string", "Bet amount", "5000.0000"),
        ("Bet List", "odds", "string", "Odds", "11.0000"),
        ("Bet List", "status", "string", "Bet status", "Khong trung"),
        # Bet Order
        ("Bet Order", "platform_id_name", "string", "Platform name", "PG NEW"),
        ("Bet Order", "platform_username", "string", "Platform username", "kisski"),
        ("Bet Order", "c_name", "string", "Category name", "No hu"),
        ("Bet Order", "game_name", "string", "Game name", "Wild Bounty Showdown"),
        ("Bet Order", "bet_amount", "int", "Bet amount", "90000"),
        ("Bet Order", "turnover", "int", "Turnover", "90000"),
        ("Bet Order", "prize", "int", "Prize amount", "0"),
        ("Bet Order", "win_lose", "int", "Win/loss", "-90000"),
        # Deposit
        ("Deposit List", "amount", "string", "Deposit amount", "500000.0000"),
        ("Deposit List", "type", "string", "Transaction type", "Nap tien"),
        ("Deposit List", "status", "string", "Transaction status", "Thanh cong"),
        # Withdrawals
        ("Withdrawals", "serial_no", "string", "Withdrawal serial", "1772308224165710055721"),
        ("Withdrawals", "amount", "string", "Withdrawal amount", "1000000.0000"),
        ("Withdrawals", "bank_name", "string", "Bank name", "Vietcombank"),
        # getLottery
        ("getLottery", "seriesData", "array", "Lottery series list", "[{id, name}]"),
        ("getLottery", "lotteryData", "array", "Lottery types list", "[{id, name, series_id}]"),
    ]

    for ri, (page, field, ftype, desc, example) in enumerate(field_entries, 2):
        ws4.cell(row=ri, column=1, value=page).border = border
        ws4.cell(row=ri, column=2, value=field).border = border
        ws4.cell(row=ri, column=3, value=ftype).border = border
        ws4.cell(row=ri, column=4, value=desc).border = border
        ws4.cell(row=ri, column=5, value=example).border = border

    ws4.freeze_panes = "A2"

    output = r"c:\Users\Admin\Desktop\MAXHUB\API_Documentation.xlsx"
    wb.save(output)
    print(f"XLSX saved: {output}")
    print(f"Total API endpoints documented: {len(API_DOCS)}")
    print(f"Sheets: API Documentation, API Summary, Date Parameters, Response Fields")

if __name__ == "__main__":
    create_xlsx()
