"""
Intercept ALL API calls from the original site with full request/response details.
Strategy:
1. Navigate to each page directly
2. Capture all network requests (especially POST to .html endpoints which are the actual API)
3. Capture response bodies
4. Also look for AJAX/fetch calls in the page JS
5. Write everything to XLSX
"""
import asyncio
import json
import sys
import urllib.request
import websockets
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')

CDP_HTTP = "http://127.0.0.1:9222"
ORIGINAL_SITE = "https://a2u4k.ee88dly.com/"

MENU_PAGES = [
    {"name": "Welcome (Trang chu)", "page": "welcome"},
    {"name": "User List (Danh sach hoi vien)", "page": "user"},
    {"name": "Invite List (Ma gioi thieu)", "page": "inviteList"},
    {"name": "Report Lottery (Bao cao xo so)", "page": "reportLottery"},
    {"name": "Report Funds (Sao ke giao dich)", "page": "reportFunds"},
    {"name": "Report Third Game (Bao cao nha cung cap)", "page": "reportThirdGame"},
    {"name": "Bank List (The ngan hang)", "page": "bankList"},
    {"name": "Deposit List (Nap tien)", "page": "depositAndWithdrawal"},
    {"name": "Withdrawals Record (Lich su rut tien)", "page": "withdrawalsRecord"},
    {"name": "Withdraw (Rut tien)", "page": "withdraw"},
    {"name": "Bet List (Don cuoc)", "page": "bet"},
    {"name": "Bet Order (Don cuoc thu 3)", "page": "betOrder"},
    {"name": "Edit Password (Doi mat khau)", "page": "editPassword"},
    {"name": "Edit Fund Password (MK giao dich)", "page": "editFundPassword"},
    {"name": "Rebate Odds (Ti le hoan tra)", "page": "getRebateOddsPanel"},
]

all_api_calls = []
msg_id = 0


def next_id():
    global msg_id
    msg_id += 1
    return msg_id


async def cdp_send(ws, method, params=None):
    cid = next_id()
    msg = {"id": cid, "method": method}
    if params:
        msg["params"] = params
    await ws.send(json.dumps(msg))
    return cid


async def cdp_recv_until(ws, cid, timeout=15):
    events = []
    deadline = asyncio.get_event_loop().time() + timeout
    while asyncio.get_event_loop().time() < deadline:
        try:
            raw = await asyncio.wait_for(ws.recv(), timeout=2)
            data = json.loads(raw)
            if data.get("id") == cid:
                return data, events
            events.append(data)
        except asyncio.TimeoutError:
            continue
    return None, events


async def drain_events(ws, duration=3):
    events = []
    deadline = asyncio.get_event_loop().time() + duration
    while asyncio.get_event_loop().time() < deadline:
        try:
            raw = await asyncio.wait_for(ws.recv(), timeout=0.3)
            events.append(json.loads(raw))
        except asyncio.TimeoutError:
            continue
    return events


async def js_eval(ws, expression):
    cid = await cdp_send(ws, "Runtime.evaluate", {
        "expression": expression,
        "returnByValue": True
    })
    result, events = await cdp_recv_until(ws, cid, timeout=10)
    val = ""
    if result and "result" in result:
        val = result["result"].get("result", {}).get("value", "")
    return val, events


async def main():
    print("=== API Interceptor V2 - Full Detail Capture ===\n")

    resp = urllib.request.urlopen(f"{CDP_HTTP}/json/list")
    tabs = json.loads(resp.read().decode())

    target_tab = None
    for tab in tabs:
        if "ee88dly" in tab.get("url", ""):
            target_tab = tab
            break

    if not target_tab:
        print("Creating new tab...")
        req = urllib.request.Request(f"{CDP_HTTP}/json/new?{ORIGINAL_SITE}", method="PUT")
        resp = urllib.request.urlopen(req)
        target_tab = json.loads(resp.read().decode())

    ws_url = target_tab["webSocketDebuggerUrl"]
    print(f"Connected to: {target_tab['url']}")

    async with websockets.connect(ws_url, max_size=100 * 1024 * 1024) as ws:
        # Enable domains
        for domain in ["Network.enable", "Page.enable", "Runtime.enable"]:
            cid = await cdp_send(ws, domain, {"maxTotalBufferSize": 100_000_000} if "Network" in domain else None)
            await cdp_recv_until(ws, cid)

        # First: go to main page to ensure session is active
        print(f"\nChecking session at {ORIGINAL_SITE}...")
        cid = await cdp_send(ws, "Page.navigate", {"url": ORIGINAL_SITE})
        await cdp_recv_until(ws, cid)
        await asyncio.sleep(5)
        await drain_events(ws, 3)

        title, _ = await js_eval(ws, "document.title")
        print(f"Page title: {title}")

        # =====================================================
        # PHASE 1: Navigate to each page and capture ALL traffic
        # =====================================================
        for menu in MENU_PAGES:
            page_name = menu["name"]
            page_key = menu["page"]
            page_url = f"{ORIGINAL_SITE}agent/{page_key}.html"

            print(f"\n{'='*60}")
            print(f"PAGE: {page_name}")
            print(f"URL: {page_url}")
            print(f"{'='*60}")

            # Track requests for this page
            pending_requests = {}  # requestId -> data
            completed_requests = {}  # requestId -> data with response

            # Navigate
            cid = await cdp_send(ws, "Page.navigate", {"url": page_url})
            result, nav_events = await cdp_recv_until(ws, cid)

            # Wait for page load
            await asyncio.sleep(4)
            all_events = nav_events + await drain_events(ws, 6)

            # Process all events
            for ev in all_events:
                m = ev.get("method", "")
                p = ev.get("params", {})

                if m == "Network.requestWillBeSent":
                    req = p.get("request", {})
                    rid = p.get("requestId", "")
                    url = req.get("url", "")
                    rtype = p.get("type", "")

                    # Skip static resources
                    skip_exts = (".css", ".js", ".png", ".jpg", ".gif", ".ico",
                                 ".svg", ".woff", ".woff2", ".ttf", ".eot", ".map")
                    if any(url.lower().endswith(ext) for ext in skip_exts):
                        continue
                    # Skip CDN challenge / RUM
                    if "/cdn-cgi/" in url:
                        continue
                    # Skip favicon
                    if "favicon" in url:
                        continue

                    pending_requests[rid] = {
                        "page": page_name,
                        "url": url,
                        "method": req.get("method", "GET"),
                        "headers": req.get("headers", {}),
                        "postData": req.get("postData", ""),
                        "type": rtype,
                        "response_status": "",
                        "response_mime": "",
                        "response_body": "",
                        "response_headers": {},
                    }

                elif m == "Network.responseReceived":
                    rid = p.get("requestId", "")
                    resp = p.get("response", {})
                    if rid in pending_requests:
                        pending_requests[rid]["response_status"] = str(resp.get("status", ""))
                        pending_requests[rid]["response_mime"] = resp.get("mimeType", "")
                        pending_requests[rid]["response_headers"] = resp.get("headers", {})

                elif m == "Network.loadingFinished":
                    rid = p.get("requestId", "")
                    if rid in pending_requests:
                        completed_requests[rid] = pending_requests.pop(rid)

            # Get response bodies for completed requests
            for rid, data in completed_requests.items():
                try:
                    cid = await cdp_send(ws, "Network.getResponseBody", {"requestId": rid})
                    result, extra = await cdp_recv_until(ws, cid, timeout=5)
                    if result and "result" in result:
                        body = result["result"].get("body", "")
                        # For HTML responses, try to extract the JSON data embedded in the page
                        if data["response_mime"] == "text/html" and body:
                            # Many layui pages embed data in script tags or layui table renders
                            # Look for JSON data patterns
                            import re
                            # Look for table data or result data
                            json_patterns = [
                                r'"data"\s*:\s*(\[.*?\])',
                                r'"rows"\s*:\s*(\[.*?\])',
                                r'"result"\s*:\s*(\{.*?\})',
                                r'var\s+\w+\s*=\s*(\{[^;]+\})\s*;',
                                r'layui\.data\([^)]*,\s*(\{[^}]+\})',
                            ]
                            extracted_json = []
                            for pat in json_patterns:
                                matches = re.findall(pat, body[:10000], re.DOTALL)
                                extracted_json.extend(matches)

                            if extracted_json:
                                data["response_body"] = "\n---\n".join(
                                    s[:2000] for s in extracted_json[:5]
                                )
                            else:
                                # Just store first portion of HTML
                                data["response_body"] = body[:3000]
                        else:
                            data["response_body"] = body[:5000]
                except Exception:
                    pass

            # Merge remaining pending into completed
            for rid, data in pending_requests.items():
                completed_requests[rid] = data

            # Print and save results
            for rid, data in completed_requests.items():
                all_api_calls.append(data)
                body_preview = data["response_body"][:100].replace("\n", " ") if data["response_body"] else ""
                print(f"  {data['method']:4s} {data['url']}")
                print(f"       Status: {data['response_status']} | Type: {data['response_mime']}")
                if data["postData"]:
                    print(f"       POST: {data['postData'][:200]}")
                if body_preview:
                    print(f"       Body: {body_preview}...")

            # =====================================================
            # Now click search/submit buttons to trigger more APIs
            # =====================================================
            print(f"\n  --- Triggering search/buttons ---")

            # Inject XHR interceptor to capture AJAX calls
            intercept_js = r"""
            (function() {
                if (window.__apiLog) return 'already injected';
                window.__apiLog = [];

                // Intercept XMLHttpRequest
                var origOpen = XMLHttpRequest.prototype.open;
                var origSend = XMLHttpRequest.prototype.send;

                XMLHttpRequest.prototype.open = function(method, url) {
                    this.__method = method;
                    this.__url = url;
                    return origOpen.apply(this, arguments);
                };

                XMLHttpRequest.prototype.send = function(body) {
                    var xhr = this;
                    var entry = {
                        method: xhr.__method,
                        url: xhr.__url,
                        requestBody: body ? String(body).substring(0, 2000) : '',
                        status: '',
                        responseBody: ''
                    };

                    xhr.addEventListener('load', function() {
                        entry.status = String(xhr.status);
                        try {
                            entry.responseBody = xhr.responseText.substring(0, 3000);
                        } catch(e) {}
                        window.__apiLog.push(entry);
                    });

                    return origSend.apply(this, arguments);
                };

                // Intercept fetch
                var origFetch = window.fetch;
                window.fetch = function(input, init) {
                    var url = typeof input === 'string' ? input : input.url;
                    var method = (init && init.method) || 'GET';
                    var body = (init && init.body) || '';

                    var entry = {
                        method: method,
                        url: url,
                        requestBody: body ? String(body).substring(0, 2000) : '',
                        status: '',
                        responseBody: ''
                    };

                    return origFetch.apply(this, arguments).then(function(response) {
                        entry.status = String(response.status);
                        return response.clone().text().then(function(text) {
                            entry.responseBody = text.substring(0, 3000);
                            window.__apiLog.push(entry);
                            return response;
                        });
                    });
                };

                // Also intercept jQuery AJAX if present
                if (window.$ && window.$.ajax) {
                    var origAjax = window.$.ajax;
                    window.$.ajax = function(options) {
                        var entry = {
                            method: (options.type || options.method || 'GET').toUpperCase(),
                            url: options.url || '',
                            requestBody: JSON.stringify(options.data || '').substring(0, 2000),
                            status: '',
                            responseBody: ''
                        };

                        var origSuccess = options.success;
                        options.success = function(data, textStatus, jqXHR) {
                            entry.status = String(jqXHR.status);
                            entry.responseBody = (typeof data === 'string' ? data : JSON.stringify(data)).substring(0, 3000);
                            window.__apiLog.push(entry);
                            if (origSuccess) origSuccess.apply(this, arguments);
                        };

                        return origAjax.call(this, options);
                    };

                    // Also intercept $.post and $.get
                    if (window.$.post) {
                        var origPost = window.$.post;
                        window.$.post = function() {
                            var args = Array.prototype.slice.call(arguments);
                            var url = args[0];
                            var data = args[1];
                            var entry = {
                                method: 'POST',
                                url: url,
                                requestBody: (typeof data === 'string' ? data : JSON.stringify(data || '')).substring(0, 2000),
                                status: '',
                                responseBody: ''
                            };

                            var result = origPost.apply(this, arguments);
                            if (result && result.done) {
                                result.done(function(resp) {
                                    entry.status = '200';
                                    entry.responseBody = (typeof resp === 'string' ? resp : JSON.stringify(resp)).substring(0, 3000);
                                    window.__apiLog.push(entry);
                                });
                            }
                            return result;
                        };
                    }
                    if (window.$.get) {
                        var origGet = window.$.get;
                        window.$.get = function() {
                            var args = Array.prototype.slice.call(arguments);
                            var url = args[0];
                            var entry = {
                                method: 'GET',
                                url: url,
                                requestBody: '',
                                status: '',
                                responseBody: ''
                            };

                            var result = origGet.apply(this, arguments);
                            if (result && result.done) {
                                result.done(function(resp) {
                                    entry.status = '200';
                                    entry.responseBody = (typeof resp === 'string' ? resp : JSON.stringify(resp)).substring(0, 3000);
                                    window.__apiLog.push(entry);
                                });
                            }
                            return result;
                        };
                    }
                }

                return 'interceptors installed (XHR + fetch + jQuery)';
            })()
            """

            inject_result, _ = await js_eval(ws, intercept_js)
            print(f"  Interceptor: {inject_result}")

            # Click search/submit buttons
            click_js = r"""
            (function() {
                var clicked = [];
                // Find and click layui search/submit buttons
                var btns = document.querySelectorAll('.layui-btn, [lay-submit], [lay-filter]');
                for (var i = 0; i < btns.length; i++) {
                    var text = btns[i].textContent.trim();
                    // Click search buttons, not dangerous ones like delete
                    if (text.match(/[Tt]ìm|[Ss]earch|查询|搜索/) || btns[i].hasAttribute('lay-submit')) {
                        btns[i].click();
                        clicked.push(text || '[submit]');
                    }
                }

                // Also trigger layui table reload if present
                if (window.layui && window.layui.table) {
                    try {
                        window.layui.table.reload('dataTable');
                        clicked.push('table.reload(dataTable)');
                    } catch(e) {}
                }

                return clicked.join(', ') || 'no buttons found';
            })()
            """

            click_result, click_events = await js_eval(ws, click_js)
            print(f"  Clicked: {click_result}")

            # Wait for AJAX calls
            await asyncio.sleep(4)
            await drain_events(ws, 3)

            # Read intercepted API log
            log_result, _ = await js_eval(ws, "JSON.stringify(window.__apiLog || [])")
            try:
                api_log = json.loads(log_result) if log_result else []
            except (json.JSONDecodeError, TypeError):
                api_log = []

            if api_log:
                print(f"  Intercepted {len(api_log)} AJAX calls:")
                for entry in api_log:
                    # Skip CDN / RUM
                    if "/cdn-cgi/" in entry.get("url", ""):
                        continue
                    print(f"    {entry['method']} {entry['url']}")
                    if entry.get("requestBody"):
                        print(f"      Request: {entry['requestBody'][:200]}")
                    if entry.get("responseBody"):
                        print(f"      Response: {entry['responseBody'][:200]}")

                    all_api_calls.append({
                        "page": f"{page_name} (AJAX)",
                        "url": entry.get("url", ""),
                        "method": entry.get("method", ""),
                        "headers": {},
                        "postData": entry.get("requestBody", ""),
                        "type": "XHR/Fetch (intercepted)",
                        "response_status": entry.get("status", ""),
                        "response_mime": "application/json",
                        "response_body": entry.get("responseBody", ""),
                        "response_headers": {},
                    })

            # =====================================================
            # Also extract API URLs from page source code
            # =====================================================
            source_js = r"""
            (function() {
                var results = [];
                var scripts = document.querySelectorAll('script');
                for (var i = 0; i < scripts.length; i++) {
                    var text = scripts[i].textContent || '';
                    if (!text) continue;

                    // Find url: patterns (jQuery AJAX)
                    var re = /url\s*:\s*['"]([^'"]+)['"]/g;
                    var m;
                    while ((m = re.exec(text)) !== null) {
                        var u = m[1];
                        if (u.indexOf('/') === 0 || u.indexOf('http') === 0) {
                            results.push({pattern: 'url:', value: u, context: text.substring(Math.max(0, m.index - 50), m.index + 100).replace(/\s+/g, ' ')});
                        }
                    }

                    // Find $.post / $.get patterns
                    re = /\$\.(post|get|ajax)\s*\(\s*['"]([^'"]+)['"]/g;
                    while ((m = re.exec(text)) !== null) {
                        results.push({pattern: '$.' + m[1], value: m[2], context: text.substring(Math.max(0, m.index - 30), m.index + 150).replace(/\s+/g, ' ')});
                    }

                    // Find type/method and data patterns near url patterns
                    re = /type\s*:\s*['"](\w+)['"]/g;
                    while ((m = re.exec(text)) !== null) {
                        results.push({pattern: 'type:', value: m[1], context: text.substring(Math.max(0, m.index - 100), m.index + 50).replace(/\s+/g, ' ')});
                    }

                    // Find data: { ... } patterns
                    re = /data\s*:\s*\{([^}]+)\}/g;
                    while ((m = re.exec(text)) !== null) {
                        results.push({pattern: 'data:', value: '{' + m[1] + '}', context: text.substring(Math.max(0, m.index - 50), m.index + 200).replace(/\s+/g, ' ')});
                    }
                }
                return JSON.stringify(results);
            })()
            """

            source_result, _ = await js_eval(ws, source_js)
            try:
                source_apis = json.loads(source_result) if source_result else []
            except (json.JSONDecodeError, TypeError):
                source_apis = []

            if source_apis:
                print(f"\n  Source code analysis ({len(source_apis)} patterns):")
                for item in source_apis:
                    print(f"    [{item['pattern']}] {item['value']}")
                    if item.get("context"):
                        print(f"      Context: {item['context'][:150]}")

                    if item["pattern"] in ("url:", "$.post", "$.get", "$.ajax"):
                        url_val = item["value"]
                        method_val = "POST" if item["pattern"] == "$.post" else \
                                     "GET" if item["pattern"] == "$.get" else "POST"

                        all_api_calls.append({
                            "page": f"{page_name} (source)",
                            "url": url_val,
                            "method": method_val,
                            "headers": {},
                            "postData": "",
                            "type": "JS_SOURCE",
                            "response_status": "N/A",
                            "response_mime": "",
                            "response_body": f"Context: {item.get('context', '')[:500]}",
                            "response_headers": {},
                        })

            # Clear interceptor log for next page
            await js_eval(ws, "window.__apiLog = [];")

        # =====================================================
        # PHASE 2: Go back to main page and click each sidebar menu item
        # to trigger the main page's iframe-based navigation
        # =====================================================
        print(f"\n\n{'='*60}")
        print("PHASE 2: Main page sidebar navigation + API capture")
        print(f"{'='*60}")

        cid = await cdp_send(ws, "Page.navigate", {"url": ORIGINAL_SITE})
        await cdp_recv_until(ws, cid)
        await asyncio.sleep(5)
        await drain_events(ws, 3)

        # Inject interceptor on main page
        await js_eval(ws, intercept_js.replace("if (window.__apiLog) return 'already injected';", "window.__apiLog = [];"))

        # Get all sidebar menu links
        menu_links_js = r"""
        (function() {
            var links = [];
            var items = document.querySelectorAll('.layui-side-scroll a[lay-href], .layui-nav-tree a[lay-href]');
            for (var i = 0; i < items.length; i++) {
                links.push({
                    text: items[i].textContent.trim(),
                    href: items[i].getAttribute('lay-href') || items[i].getAttribute('href') || '',
                    index: i
                });
            }
            // Also try dd a tags in sidebar
            var dd_links = document.querySelectorAll('.layui-side dd a');
            for (var i = 0; i < dd_links.length; i++) {
                var h = dd_links[i].getAttribute('lay-href') || dd_links[i].getAttribute('href') || '';
                if (h && h !== 'javascript:;') {
                    links.push({text: dd_links[i].textContent.trim(), href: h, index: 100 + i});
                }
            }
            return JSON.stringify(links);
        })()
        """
        menu_result, _ = await js_eval(ws, menu_links_js)
        try:
            sidebar_links = json.loads(menu_result) if menu_result else []
        except (json.JSONDecodeError, TypeError):
            sidebar_links = []

        print(f"Found {len(sidebar_links)} sidebar links:")
        for link in sidebar_links:
            print(f"  {link['text']}: {link['href']}")

        # Click each sidebar link and capture iframe API calls
        for link in sidebar_links:
            if not link.get("href") or link["href"] == "javascript:;":
                continue

            print(f"\n  Clicking: {link['text']} -> {link['href']}")

            # Clear interceptor
            await js_eval(ws, "window.__apiLog = [];")

            # Click the link
            click_link_js = f"""
            (function() {{
                var links = document.querySelectorAll('.layui-side-scroll a, .layui-nav-tree a, .layui-side dd a');
                for (var i = 0; i < links.length; i++) {{
                    var h = links[i].getAttribute('lay-href') || links[i].getAttribute('href') || '';
                    if (h === '{link["href"]}') {{
                        links[i].click();
                        return 'clicked';
                    }}
                }}
                return 'not found';
            }})()
            """
            result, _ = await js_eval(ws, click_link_js)
            print(f"    Click: {result}")

            await asyncio.sleep(3)
            events = await drain_events(ws, 4)

            # Check for API calls from network events
            for ev in events:
                m = ev.get("method", "")
                p = ev.get("params", {})
                if m == "Network.requestWillBeSent":
                    req = p.get("request", {})
                    url = req.get("url", "")
                    if "/cdn-cgi/" in url:
                        continue
                    skip_exts = (".css", ".js", ".png", ".jpg", ".gif", ".ico",
                                 ".svg", ".woff", ".woff2", ".ttf", ".eot")
                    if any(url.lower().endswith(ext) for ext in skip_exts):
                        continue
                    rtype = p.get("type", "")
                    if rtype in ("XHR", "Fetch") or req.get("method") == "POST":
                        print(f"    Network: {req.get('method')} {url}")

            # Check interceptor log
            log_result, _ = await js_eval(ws, "JSON.stringify(window.__apiLog || [])")
            try:
                log = json.loads(log_result) if log_result else []
            except (json.JSONDecodeError, TypeError):
                log = []

            for entry in log:
                if "/cdn-cgi/" in entry.get("url", ""):
                    continue
                print(f"    AJAX: {entry['method']} {entry['url']}")
                all_api_calls.append({
                    "page": f"Sidebar: {link['text']}",
                    "url": entry.get("url", ""),
                    "method": entry.get("method", ""),
                    "headers": {},
                    "postData": entry.get("requestBody", ""),
                    "type": "XHR/Fetch (sidebar)",
                    "response_status": entry.get("status", ""),
                    "response_mime": "",
                    "response_body": entry.get("responseBody", ""),
                    "response_headers": {},
                })

        # =====================================================
        # PHASE 3: Extract getLottery and other known API patterns
        # =====================================================
        print(f"\n\n{'='*60}")
        print("PHASE 3: Known API endpoint testing")
        print(f"{'='*60}")

        # From Phase 1 we saw /agent/getLottery being called
        # Let's directly test known API patterns
        known_endpoints = [
            {"url": "/agent/welcome", "method": "POST", "data": ""},
            {"url": "/agent/user", "method": "POST", "data": "page=1&limit=10"},
            {"url": "/agent/inviteList", "method": "POST", "data": "page=1&limit=10"},
            {"url": "/agent/reportLottery", "method": "POST", "data": "page=1&limit=10&startDate=2026-03-01&endDate=2026-03-01"},
            {"url": "/agent/reportFunds", "method": "POST", "data": "page=1&limit=10&startDate=2026-03-01&endDate=2026-03-01"},
            {"url": "/agent/reportThirdGame", "method": "POST", "data": "page=1&limit=10&startDate=2026-03-01&endDate=2026-03-01"},
            {"url": "/agent/bankList", "method": "POST", "data": "page=1&limit=10"},
            {"url": "/agent/depositAndWithdrawal", "method": "POST", "data": "page=1&limit=10"},
            {"url": "/agent/withdrawalsRecord", "method": "POST", "data": "page=1&limit=10"},
            {"url": "/agent/bet", "method": "POST", "data": "page=1&limit=10&startDate=2026-03-01&endDate=2026-03-01"},
            {"url": "/agent/betOrder", "method": "POST", "data": "page=1&limit=10&startDate=2026-03-01&endDate=2026-03-01"},
            {"url": "/agent/getLottery", "method": "POST", "data": ""},
            {"url": "/agent/getRebateOddsPanel", "method": "POST", "data": ""},
        ]

        for ep in known_endpoints:
            fetch_js = f"""
            (function() {{
                return new Promise(function(resolve) {{
                    var xhr = new XMLHttpRequest();
                    xhr.open('{ep["method"]}', '{ep["url"]}', true);
                    xhr.setRequestHeader('Content-Type', 'application/x-www-form-urlencoded');
                    xhr.setRequestHeader('X-Requested-With', 'XMLHttpRequest');
                    xhr.onload = function() {{
                        resolve(JSON.stringify({{
                            status: xhr.status,
                            headers: xhr.getAllResponseHeaders(),
                            body: xhr.responseText.substring(0, 5000)
                        }}));
                    }};
                    xhr.onerror = function() {{
                        resolve(JSON.stringify({{status: 0, error: 'network error'}}));
                    }};
                    xhr.send('{ep["data"]}');
                }});
            }})()
            """

            cid = await cdp_send(ws, "Runtime.evaluate", {
                "expression": fetch_js,
                "awaitPromise": True,
                "returnByValue": True,
            })
            result, _ = await cdp_recv_until(ws, cid, timeout=15)

            resp_str = result.get("result", {}).get("result", {}).get("value", "{}") if result else "{}"
            try:
                resp_data = json.loads(resp_str)
            except (json.JSONDecodeError, TypeError):
                resp_data = {"status": "error", "body": resp_str}

            status = resp_data.get("status", "?")
            body = resp_data.get("body", "")
            print(f"\n  {ep['method']} {ep['url']} -> {status}")
            if ep.get("data"):
                print(f"    Params: {ep['data']}")
            if body:
                print(f"    Response: {body[:300]}")

            all_api_calls.append({
                "page": "Direct API Test",
                "url": ep["url"],
                "method": ep["method"],
                "headers": {"Content-Type": "application/x-www-form-urlencoded", "X-Requested-With": "XMLHttpRequest"},
                "postData": ep.get("data", ""),
                "type": "Direct XHR",
                "response_status": str(status),
                "response_mime": "text/html or application/json",
                "response_body": body[:5000],
                "response_headers": {},
            })

    print(f"\n\n{'='*60}")
    print(f"TOTAL API CALLS CAPTURED: {len(all_api_calls)}")
    print(f"{'='*60}")

    write_xlsx(all_api_calls)


def write_xlsx(api_calls):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "API Documentation"

    hfont = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        "STT", "Page/Menu", "API Endpoint", "HTTP Method", "Source Type",
        "Request Content-Type", "Request Params/Body",
        "Response Status", "Response Content-Type",
        "Response Body (sample)", "Notes"
    ]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = halign
        c.border = border

    widths = [6, 32, 60, 12, 20, 30, 60, 14, 22, 100, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Deduplicate by url+method (keep first occurrence with most data)
    seen = {}
    unique = []
    for c in api_calls:
        key = f"{c['url']}|{c['method']}"
        if key not in seen:
            seen[key] = len(unique)
            unique.append(c)
        else:
            # Keep the one with more response data
            existing = unique[seen[key]]
            if len(c.get("response_body", "")) > len(existing.get("response_body", "")):
                unique[seen[key]] = c

    dfont = Font(name="Arial", size=10)
    walign = Alignment(vertical="top", wrap_text=True)

    for ri, call in enumerate(unique, 2):
        # Format content-type from headers
        ct = ""
        hdrs = call.get("headers", {})
        if isinstance(hdrs, dict):
            ct = hdrs.get("Content-Type", hdrs.get("content-type", ""))

        post_data = call.get("postData", "")
        if post_data:
            try:
                post_data = json.dumps(json.loads(post_data), indent=2, ensure_ascii=False)
            except (json.JSONDecodeError, TypeError):
                pass

        resp_body = call.get("response_body", "")
        if resp_body:
            try:
                parsed = json.loads(resp_body)
                resp_body = json.dumps(parsed, indent=2, ensure_ascii=False)[:4000]
            except (json.JSONDecodeError, TypeError):
                resp_body = resp_body[:4000]

        row = [
            ri - 1,
            call.get("page", ""),
            call.get("url", ""),
            call.get("method", ""),
            call.get("type", ""),
            ct,
            post_data,
            str(call.get("response_status", "")),
            call.get("response_mime", ""),
            resp_body,
            "",
        ]

        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=str(val))
            c.font = dfont
            c.alignment = walign
            c.border = border

    # Summary sheet
    ws2 = wb.create_sheet("Summary by Page")
    for ci, h in enumerate(["Page", "API Count", "Endpoints"], 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = hfont
        c.fill = hfill
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 90

    page_map = {}
    for c in unique:
        pg = c.get("page", "?")
        page_map.setdefault(pg, []).append(f"{c['method']} {c['url']}")

    for ri, (pg, urls) in enumerate(page_map.items(), 2):
        ws2.cell(row=ri, column=1, value=pg)
        ws2.cell(row=ri, column=2, value=len(urls))
        ws2.cell(row=ri, column=3, value="\n".join(urls))
        ws2.cell(row=ri, column=3).alignment = walign

    ws.freeze_panes = "A2"
    ws2.freeze_panes = "A2"
    if unique:
        ws.auto_filter.ref = f"A1:K{len(unique) + 1}"

    out = r"c:\Users\Admin\Desktop\MAXHUB\API_Documentation.xlsx"
    wb.save(out)
    print(f"\nXLSX saved: {out}")
    print(f"Unique API calls: {len(unique)}")


if __name__ == "__main__":
    asyncio.run(main())
