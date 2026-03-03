#!/usr/bin/env python3
"""
Re-import dữ liệu từ file template gốc DINHDANGBIEUDOANHTHU.xlsx vào DB.
Đọc đúng cấu trúc: mỗi sheet = 1 NV, cột A = ngày, cột D = username.
Thay thế toàn bộ data cũ trong employees + employee_customers.
"""
import sys
from datetime import datetime

import openpyxl
import psycopg2

DB_URL = "postgresql://postgres:hiepmun2021@localhost:5432/fastify_skeleton"
TEMPLATE_FILE = "DINHDANGBIEUDOANHTHU.xlsx"
SKIP_SHEETS = {"TỔNG QUÁT", "CHI TIẾT DOANH THU"}


def parse_date(value):
    """Parse ngày từ cell value → 'YYYY-MM-DD' hoặc None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        # Excel serial date
        from datetime import timedelta
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=int(value))).strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s or "💰" in s or "TỔNG" in s or "LỢI NHUẬN" in s:
        return None
    # DD/MM/YYYY
    import re
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    # YYYY-MM-DD
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if m:
        return s
    return None


def extract_customers(ws, sheet_name):
    """Extract (username, assigned_date) từ 1 sheet nhân viên."""
    customers = []
    current_date = None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=5):
        a_val = row[0].value  # Cột A: ngày
        d_val = row[3].value  # Cột D: username

        # Cập nhật ngày nếu cột A có giá trị
        parsed = parse_date(a_val)
        if parsed:
            current_date = parsed

        # Kiểm tra có username không
        if not d_val:
            continue
        username = str(d_val).strip().lower()
        if not username or len(username) < 2:
            continue
        # Bỏ qua nếu là header/summary
        if any(kw in username for kw in ["客户", "tài khoản", "💰", "tổng cộng", "lợi nhuận"]):
            continue

        customers.append((username, current_date))

    return customers


def main():
    template_file = sys.argv[1] if len(sys.argv) > 1 else TEMPLATE_FILE

    print(f"Đọc file template: {template_file}")
    wb = openpyxl.load_workbook(template_file, data_only=True, read_only=True)

    # Collect all employee data
    all_employees = {}  # name → [(username, date), ...]
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        customers = extract_customers(ws, sheet_name)
        if customers:
            all_employees[sheet_name] = customers
            print(f"  {sheet_name}: {len(customers)} khách")
        else:
            print(f"  {sheet_name}: 0 khách (bỏ qua)")

    wb.close()

    total_customers = sum(len(v) for v in all_employees.values())
    print(f"\nTổng: {len(all_employees)} nhân viên, {total_customers:,} khách")

    # Import vào DB
    print("\nĐang import vào DB...")
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()

    try:
        # Xóa data cũ
        cur.execute("DELETE FROM employee_customers")
        cur.execute("DELETE FROM employees")
        deleted_ec = cur.rowcount
        print(f"  Đã xóa data cũ")

        # Insert từng nhân viên
        total_inserted = 0
        for emp_name, customers in all_employees.items():
            # Tạo employee
            cur.execute(
                "INSERT INTO employees (id, name, created_at, updated_at) VALUES (gen_random_uuid()::text, %s, NOW(), NOW()) RETURNING id",
                (emp_name,),
            )
            emp_id = cur.fetchone()[0]

            # Insert customers (deduplicate by username)
            seen = set()
            batch = []
            for username, assigned_date in customers:
                if username in seen:
                    continue
                seen.add(username)
                batch.append((emp_id, username, assigned_date))

            if batch:
                # Batch insert
                args = ",".join(
                    cur.mogrify(
                        "(gen_random_uuid()::text, %s, %s, %s, NOW())",
                        (eid, uname, adate),
                    ).decode()
                    for eid, uname, adate in batch
                )
                cur.execute(
                    f"INSERT INTO employee_customers (id, employee_id, customer_username, assigned_date, created_at) VALUES {args}"
                )
                total_inserted += len(batch)

            print(f"  {emp_name}: {len(batch)} unique / {len(customers)} total")

        conn.commit()
        print(f"\n✅ Import thành công: {len(all_employees)} NV, {total_inserted:,} khách")

    except Exception as e:
        conn.rollback()
        print(f"\n❌ Lỗi: {e}")
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
