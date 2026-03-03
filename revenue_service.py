#!/usr/bin/env python3
"""
Script xuất biểu doanh thu theo đúng template DINHDANGBIEUDOANHTHU.xlsx
Dùng: python revenue_service.py [YYYY-MM]
Mặc định: tháng hiện tại
Output: DOANH_THU_THANG_MM_YYYY.xlsx
"""
import sys
import calendar
from collections import defaultdict, OrderedDict
from datetime import datetime

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════

DB_URL = "postgresql://postgres:hiepmun2021@localhost:5432/fastify_skeleton"
NUM_FMT = "#,##0"

# ═══════════════════════════════════════════════════════════════════
# STYLES — giữ đúng template gốc
# ═══════════════════════════════════════════════════════════════════

THIN = Side(style="thin")
BORDER_ALL = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center")

FONT_BASE     = Font(name="Times New Roman", size=10)
FONT_BOLD     = Font(name="Times New Roman", size=10, bold=True)
FONT_BOLD_11  = Font(name="Times New Roman", size=11, bold=True)
FONT_BOLD_12W = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
FONT_BOLD_14  = Font(name="Times New Roman", size=14, bold=True)
FONT_BOLD_14W = Font(name="Times New Roman", size=14, bold=True, color="FFFFFF")
FONT_BOLD_16W = Font(name="Times New Roman", size=16, bold=True, color="FFFFFF")
FONT_BOLD_W   = Font(name="Times New Roman", size=10, bold=True, color="FFFFFF")
FONT_BOLD_11W = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")

FILL_GREEN    = PatternFill("solid", fgColor="2E7D32")   # TỔNG QUÁT header
FILL_GOLD     = PatternFill("solid", fgColor="FFD700")   # TỔNG QUÁT data
FILL_RED      = PatternFill("solid", fgColor="FF0000")   # TỔNG QUÁT total
FILL_DKBLUE   = PatternFill("solid", fgColor="1F4E79")   # CHI TIẾT title
FILL_BLUE     = PatternFill("solid", fgColor="4472C4")   # header rows
FILL_LTBLUE   = PatternFill("solid", fgColor="D9E2F3")   # month label
FILL_DKRED    = PatternFill("solid", fgColor="C00000")   # CHI TIẾT total
FILL_YELLOW   = PatternFill("solid", fgColor="FFFF00")   # employee month


# ═══════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════

def style_row(ws, row_num, font, fill, alignment, border, num_fmt=None,
              start_col=1, end_col=None):
    """Áp style cho toàn bộ row từ start_col đến end_col."""
    if end_col is None:
        end_col = ws.max_column or start_col
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row_num, column=c)
        if font:  cell.font = font
        if fill:  cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = border
        if num_fmt and c >= 3:  # chỉ áp number format cho cột số
            cell.number_format = num_fmt


def fmt_month_label(yyyy_mm):
    """'2026-02' → 'THÁNG 2/2026'"""
    if "-" not in yyyy_mm:
        return yyyy_mm
    y, m = yyyy_mm.split("-")
    return f"THÁNG {int(m)}/{y}"


def fmt_month_summary(yyyy_mm):
    """'2026-02' → '💰 LỢI NHUẬN THÁNG 02/2026'"""
    if "-" not in yyyy_mm:
        return f"💰 LỢI NHUẬN ({yyyy_mm})"
    y, m = yyyy_mm.split("-")
    return f"💰 LỢI NHUẬN THÁNG {m}/{y}"


def fmt_date_vn(date_str):
    """'2026-02-15' → '15/02/2026'"""
    if not date_str:
        return ""
    parts = date_str.split("-")
    if len(parts) == 3:
        return f"{parts[2]}/{parts[1]}/{parts[0]}"
    return date_str


# ═══════════════════════════════════════════════════════════════════
# DATA FETCHING
# ═══════════════════════════════════════════════════════════════════

def fetch_data(month):
    """Lấy toàn bộ dữ liệu cần thiết cho biểu doanh thu."""
    y, m = map(int, month.split("-"))
    start_date = f"{month}-01"
    last_day = calendar.monthrange(y, m)[1]
    end_date = f"{month}-{last_day:02d}"

    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()

    # 1. Employees
    cur.execute("SELECT id, name FROM employees ORDER BY name")
    employees = {row[0]: row[1] for row in cur.fetchall()}

    # 2. Employee-Customer mappings
    cur.execute("""
        SELECT employee_id, customer_username, assigned_date
        FROM employee_customers
    """)
    emp_customers = defaultdict(list)
    for eid, username, assigned_date in cur.fetchall():
        emp_customers[eid].append({
            "username": username,
            "assigned_date": str(assigned_date) if assigned_date else None,
        })

    # 3. Revenue per employee for selected month
    # Lottery
    cur.execute("""
        SELECT ec.employee_id, COALESCE(SUM(r.win_lose), 0) AS total
        FROM employee_customers ec
        JOIN proxy_report_lottery r ON r.username = ec.customer_username
        WHERE r.report_date >= %s AND r.report_date <= %s
        GROUP BY ec.employee_id
    """, (start_date, end_date))
    lottery_by_emp = dict(cur.fetchall())

    # Third game
    cur.execute("""
        SELECT ec.employee_id, COALESCE(SUM(r.t_win_lose), 0) AS total
        FROM employee_customers ec
        JOIN proxy_report_third_game r ON r.username = ec.customer_username
        WHERE r.report_date >= %s AND r.report_date <= %s
        GROUP BY ec.employee_id
    """, (start_date, end_date))
    third_by_emp = dict(cur.fetchall())

    # Funds (promotion + third_rebate)
    cur.execute("""
        SELECT ec.employee_id,
               COALESCE(SUM(r.promotion), 0),
               COALESCE(SUM(r.third_rebate), 0)
        FROM employee_customers ec
        JOIN proxy_report_funds r ON r.username = ec.customer_username
        WHERE r.report_date >= %s AND r.report_date <= %s
        GROUP BY ec.employee_id
    """, (start_date, end_date))
    funds_by_emp = {}
    for eid, promo, rebate in cur.fetchall():
        funds_by_emp[eid] = (float(promo), float(rebate))

    # 4. Per-username revenue for employee detail sheets (selected month only)
    cur.execute("""
        SELECT ec.customer_username, COALESCE(SUM(r.win_lose), 0)
        FROM employee_customers ec
        JOIN proxy_report_lottery r ON r.username = ec.customer_username
        WHERE r.report_date >= %s AND r.report_date <= %s
        GROUP BY ec.customer_username
    """, (start_date, end_date))
    lottery_by_user = dict(cur.fetchall())

    cur.execute("""
        SELECT ec.customer_username, COALESCE(SUM(r.t_win_lose), 0)
        FROM employee_customers ec
        JOIN proxy_report_third_game r ON r.username = ec.customer_username
        WHERE r.report_date >= %s AND r.report_date <= %s
        GROUP BY ec.customer_username
    """, (start_date, end_date))
    third_by_user = dict(cur.fetchall())

    cur.execute("""
        SELECT ec.customer_username,
               COALESCE(SUM(r.promotion), 0),
               COALESCE(SUM(r.third_rebate), 0)
        FROM employee_customers ec
        JOIN proxy_report_funds r ON r.username = ec.customer_username
        WHERE r.report_date >= %s AND r.report_date <= %s
        GROUP BY ec.customer_username
    """, (start_date, end_date))
    funds_by_user = {}
    for username, promo, rebate in cur.fetchall():
        funds_by_user[username] = (float(promo), float(rebate))

    # 5. Revenue matrix — all months, per employee
    cur.execute("""
        SELECT ec.employee_id,
               TO_CHAR(r.report_date::date, 'YYYY-MM') AS month,
               SUM(r.total)::float AS total
        FROM employee_customers ec
        JOIN (
            SELECT username, report_date, win_lose AS total FROM proxy_report_lottery
            UNION ALL
            SELECT username, report_date, t_win_lose AS total FROM proxy_report_third_game
            UNION ALL
            SELECT username, report_date, promotion AS total FROM proxy_report_funds
            UNION ALL
            SELECT username, report_date, third_rebate AS total FROM proxy_report_funds
        ) r ON r.username = ec.customer_username
        GROUP BY ec.employee_id, TO_CHAR(r.report_date::date, 'YYYY-MM')
        ORDER BY month
    """)
    matrix_data = defaultdict(lambda: defaultdict(float))
    all_months = set()
    for eid, mth, total in cur.fetchall():
        matrix_data[mth][eid] += float(total)
        all_months.add(mth)

    conn.close()

    # Build employee summaries
    emp_summaries = []
    for eid, name in employees.items():
        lottery = float(lottery_by_emp.get(eid, 0))
        third = float(third_by_emp.get(eid, 0))
        promo, rebate = funds_by_emp.get(eid, (0.0, 0.0))
        total_rev = lottery + third + promo + rebate
        emp_summaries.append({
            "id": eid,
            "name": name,
            "lottery": lottery,
            "third": third,
            "promotion": promo,
            "rebate": rebate,
            "total": total_rev,
            "customer_count": len(emp_customers.get(eid, [])),
        })

    # Sort by total ascending (most negative = most profit first)
    emp_summaries.sort(key=lambda e: e["total"])

    return {
        "month": month,
        "employees": employees,
        "emp_summaries": emp_summaries,
        "emp_customers": emp_customers,
        "lottery_by_user": lottery_by_user,
        "third_by_user": third_by_user,
        "funds_by_user": funds_by_user,
        "matrix_data": matrix_data,
        "all_months": sorted(all_months),
    }


# ═══════════════════════════════════════════════════════════════════
# SHEET 1: TỔNG QUÁT
# ═══════════════════════════════════════════════════════════════════

def build_tong_quat(wb, data):
    ws = wb.create_sheet("TỔNG QUÁT", 0)

    # Column widths
    for col, w in enumerate([8, 22, 18, 20, 14, 18, 20], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Header
    headers = [
        "序号\nSTT",
        "员工姓名\nTên nhân viên",
        "彩票利润\nLợi nhuận XS",
        "第三者利润\nLợi nhuận bên thứ 3",
        "优惠\nƯu đãi",
        "第三者退款\nHoàn trả",
        "总营收\nTổng doanh thu",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
    ws.row_dimensions[1].height = 40
    style_row(ws, 1, FONT_BOLD_W, FILL_GREEN, ALIGN_CENTER, BORDER_ALL)

    # Data rows
    gt = {"lottery": 0, "third": 0, "promo": 0, "rebate": 0, "total": 0}
    for i, emp in enumerate(data["emp_summaries"]):
        r = i + 2
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=emp["name"])
        ws.cell(row=r, column=3, value=emp["lottery"])
        ws.cell(row=r, column=4, value=emp["third"])
        ws.cell(row=r, column=5, value=emp["promotion"])
        ws.cell(row=r, column=6, value=emp["rebate"])
        ws.cell(row=r, column=7, value=emp["total"])
        style_row(ws, r, FONT_BOLD_11, FILL_GOLD, ALIGN_CENTER, BORDER_ALL, NUM_FMT)
        gt["lottery"] += emp["lottery"]
        gt["third"] += emp["third"]
        gt["promo"] += emp["promotion"]
        gt["rebate"] += emp["rebate"]
        gt["total"] += emp["total"]

    # Total row
    tr = len(data["emp_summaries"]) + 2
    ws.cell(row=tr, column=1, value="总计\nTỔNG CỘNG")
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=2)
    ws.cell(row=tr, column=3, value=gt["lottery"])
    ws.cell(row=tr, column=4, value=gt["third"])
    ws.cell(row=tr, column=5, value=gt["promo"])
    ws.cell(row=tr, column=6, value=gt["rebate"])
    ws.cell(row=tr, column=7, value=gt["total"])
    style_row(ws, tr, FONT_BOLD_14W, FILL_RED, ALIGN_CENTER, BORDER_ALL, NUM_FMT)


# ═══════════════════════════════════════════════════════════════════
# SHEET 2: CHI TIẾT DOANH THU
# ═══════════════════════════════════════════════════════════════════

def build_chi_tiet(wb, data):
    ws = wb.create_sheet("CHI TIẾT DOANH THU", 1)
    emp_list = data["emp_summaries"]
    n_emp = len(emp_list)
    total_cols = n_emp + 2  # A + employees + TỔNG

    # Column widths
    ws.column_dimensions["A"].width = 16
    for i in range(n_emp):
        ws.column_dimensions[get_column_letter(i + 2)].width = 15
    ws.column_dimensions[get_column_letter(total_cols)].width = 18

    # Row 1: Title
    ws.cell(row=1, column=1, value="各月利润明细\nLỢI NHUẬN CHI TIẾT THEO THÁNG")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.row_dimensions[1].height = 35
    style_row(ws, 1, FONT_BOLD_16W, FILL_DKBLUE, ALIGN_CENTER, BORDER_ALL,
              end_col=total_cols)

    # Row 2: Employee headers
    ws.cell(row=2, column=1, value="开发人\nNHÂN VIÊN")
    for i, emp in enumerate(emp_list):
        ws.cell(row=2, column=i + 2, value=emp["name"])
    ws.cell(row=2, column=total_cols, value="总营收\nTỔNG")
    ws.row_dimensions[2].height = 30
    style_row(ws, 2, FONT_BOLD_11W, FILL_BLUE, ALIGN_CENTER, BORDER_ALL,
              end_col=total_cols)

    # Data rows per month
    emp_id_to_col = {emp["id"]: i + 2 for i, emp in enumerate(emp_list)}
    r = 3
    for mth in data["all_months"]:
        ws.cell(row=r, column=1, value=fmt_month_label(mth))
        ws.cell(row=r, column=1).font = FONT_BOLD
        ws.cell(row=r, column=1).fill = FILL_LTBLUE
        ws.cell(row=r, column=1).border = BORDER_ALL

        row_total = 0
        for emp in emp_list:
            col = emp_id_to_col[emp["id"]]
            val = data["matrix_data"].get(mth, {}).get(emp["id"], 0)
            if val != 0:
                ws.cell(row=r, column=col, value=val)
                ws.cell(row=r, column=col).number_format = NUM_FMT
            ws.cell(row=r, column=col).border = BORDER_ALL
            row_total += val

        ws.cell(row=r, column=total_cols, value=row_total if row_total != 0 else None)
        ws.cell(row=r, column=total_cols).number_format = NUM_FMT
        ws.cell(row=r, column=total_cols).border = BORDER_ALL
        r += 1

    # Total row: LỢI NHUẬN
    ws.cell(row=r, column=1, value="利润\nLỢI NHUẬN")
    grand_total = 0
    for emp in emp_list:
        col = emp_id_to_col[emp["id"]]
        emp_total = sum(
            data["matrix_data"].get(m, {}).get(emp["id"], 0)
            for m in data["all_months"]
        )
        ws.cell(row=r, column=col, value=emp_total)
        grand_total += emp_total
    ws.cell(row=r, column=total_cols, value=grand_total)
    style_row(ws, r, FONT_BOLD_12W, FILL_DKRED, ALIGN_CENTER, BORDER_ALL, NUM_FMT,
              end_col=total_cols)


# ═══════════════════════════════════════════════════════════════════
# EMPLOYEE SHEETS
# ═══════════════════════════════════════════════════════════════════

def build_employee_sheet(wb, data, emp):
    eid = emp["id"]
    name = emp["name"]
    ws = wb.create_sheet(name)

    # Column widths
    col_widths = [14, 18, 22, 24, 16, 18, 20, 14, 20, 14, 18]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Header
    headers = [
        "日期\nNgày",
        "开发客户总数量\nTổng lượng khách",
        "今日开发客户数量\nLượng khách trong ngày",
        "客户账号\nTài khoản khách hàng",
        "首次充值\nNạp tiền lần đầu",
        "彩票利润\nLợi nhuận xổ số",
        "第三者利润\nLợi nhuận bên thứ 3",
        "优惠\nƯu Đãi",
        "第三者退款\nHoàn trả bên thứ 3",
        "总\nTổng",
        "输赢 2M\nTHẮNG THUA",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    ws.row_dimensions[1].height = 40
    style_row(ws, 1, FONT_BOLD_W, FILL_BLUE, ALIGN_CENTER, BORDER_ALL, end_col=11)

    # Get customers for this employee
    customers = data["emp_customers"].get(eid, [])
    if not customers:
        return

    # Sort by assigned_date
    customers.sort(key=lambda c: c["assigned_date"] or "9999-99-99")

    # Group by month
    by_month = OrderedDict()
    for c in customers:
        mkey = c["assigned_date"][:7] if c["assigned_date"] else "unknown"
        by_month.setdefault(mkey, []).append(c)

    r = 2
    running_total = 0
    month_summary_rows = []

    for mkey, month_custs in by_month.items():
        month_start_row = r

        # Group by date within month
        by_date = OrderedDict()
        for c in month_custs:
            dkey = c["assigned_date"] or "unknown"
            by_date.setdefault(dkey, []).append(c)

        for date_str, day_custs in by_date.items():
            day_start_row = r
            for i, cust in enumerate(day_custs):
                running_total += 1
                username = cust["username"]
                lottery = float(data["lottery_by_user"].get(username, 0))
                third = float(data["third_by_user"].get(username, 0))
                promo, rebate = data["funds_by_user"].get(username, (0.0, 0.0))
                total = lottery + third + promo + rebate

                if i == 0:
                    ws.cell(row=r, column=1, value=fmt_date_vn(date_str))
                    ws.cell(row=r, column=2, value=running_total)
                    ws.cell(row=r, column=3, value=len(day_custs))

                ws.cell(row=r, column=4, value=username)
                ws.cell(row=r, column=5, value=0)  # firstDeposit — chưa có data
                ws.cell(row=r, column=6, value=lottery)
                ws.cell(row=r, column=7, value=third)
                ws.cell(row=r, column=8, value=promo)
                ws.cell(row=r, column=9, value=rebate)
                ws.cell(row=r, column=10, value=total)
                # K: empty

                style_row(ws, r, FONT_BASE, None, ALIGN_CENTER, BORDER_ALL,
                          NUM_FMT, end_col=11)
                r += 1

            # Merge date/count cells if multiple customers same day
            if len(day_custs) > 1:
                for col in [1, 2, 3]:
                    ws.merge_cells(
                        start_row=day_start_row, start_column=col,
                        end_row=r - 1, end_column=col,
                    )

        # Month summary row
        month_end_row = r - 1
        ws.cell(row=r, column=1, value=fmt_month_summary(mkey))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

        # SUM formulas for F, G, H, I, J
        for col_idx, col_letter in [(6, "F"), (7, "G"), (8, "H"), (9, "I"), (10, "J")]:
            ws.cell(row=r, column=col_idx,
                    value=f"=SUM({col_letter}{month_start_row}:{col_letter}{month_end_row})")

        style_row(ws, r, FONT_BOLD_14, FILL_YELLOW, ALIGN_CENTER, BORDER_ALL,
                  NUM_FMT, end_col=11)
        month_summary_rows.append(r)
        r += 1

    # Grand total row
    ws.cell(row=r, column=1, value="TỔNG CỘNG")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

    if month_summary_rows:
        for col_idx, col_letter in [(6, "F"), (7, "G"), (8, "H"), (9, "I"), (10, "J")]:
            refs = ",".join(f"{col_letter}{sr}" for sr in month_summary_rows)
            ws.cell(row=r, column=col_idx, value=f"=SUM({refs})")

    style_row(ws, r, FONT_BOLD_14W, FILL_RED, ALIGN_CENTER, BORDER_ALL,
              NUM_FMT, end_col=11)


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

def main():
    # Parse month argument
    if len(sys.argv) > 1:
        month = sys.argv[1]
    else:
        now = datetime.now()
        month = f"{now.year}-{now.month:02d}"

    print(f"Đang lấy dữ liệu tháng {month}...")
    data = fetch_data(month)

    print(f"  → {len(data['emp_summaries'])} nhân viên")
    total_customers = sum(len(v) for v in data["emp_customers"].values())
    print(f"  → {total_customers:,} khách hàng")
    print(f"  → {len(data['all_months'])} tháng trong matrix")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Tạo sheet TỔNG QUÁT...")
    build_tong_quat(wb, data)

    print("Tạo sheet CHI TIẾT DOANH THU...")
    build_chi_tiet(wb, data)

    print("Tạo sheet cho từng nhân viên...")
    for emp in data["emp_summaries"]:
        customers = data["emp_customers"].get(emp["id"], [])
        if customers:
            print(f"  → {emp['name']} ({len(customers)} khách)")
            build_employee_sheet(wb, data, emp)
        else:
            print(f"  → {emp['name']} (bỏ qua — 0 khách)")

    # Save
    y, m = month.split("-")
    filename = f"DOANH_THU_THANG_{m}_{y}.xlsx"
    # Nếu file bị lock (đang mở trong Excel), thử tên khác
    import os
    if os.path.exists(filename):
        try:
            with open(filename, "a"):
                pass
        except PermissionError:
            filename = f"DOANH_THU_THANG_{m}_{y}_new.xlsx"
            print(f"  ⚠ File cũ đang mở, lưu thành: {filename}")
    wb.save(filename)
    print(f"\n✅ Đã xuất: {filename}")


if __name__ == "__main__":
    main()
