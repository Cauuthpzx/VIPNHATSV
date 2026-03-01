import asyncio
import json
import sys
import urllib.request
import websockets
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')

CDP_HTTP = "http://127.0.0.1:9222"
ORIGINAL_SITE = "https://a2u4k.ee88dly.com/"

MENU_PAGES = [
    {"name": "Trang chu", "page": "welcome"},
    {"name": "Danh sach hoi vien", "page": "user"},
    {"name": "Danh sach ma gioi thieu", "page": "inviteList"},
    {"name": "Bao cao xo so", "page": "reportLottery"},
    {"name": "Sao ke giao dich", "page": "reportFunds"},
    {"name": "Bao cao nha cung cap", "page": "reportThirdGame"},
    {"name": "Danh sach the ngan hang", "page": "bankList"},
    {"name": "Danh sach nap tien", "page": "depositAndWithdrawal"},
    {"name": "Lich su rut tien", "page": "withdrawalsRecord"},
    {"name": "Rut tien", "page": "withdraw"},
    {"name": "Danh sach don cuoc", "page": "bet"},
    {"name": "Don cuoc ben thu 3", "page": "betOrder"},
    {"name": "Doi mat khau dang nhap", "page": "editPassword"},
    {"name": "Doi mat khau giao dich", "page": "editFundPassword"},
    {"name": "Danh sach ti le hoan tra", "page": "getRebateOddsPanel"},
]

all_api_calls = []
msg_id = 0


def next_id():
    global msg_id
    msg_id += 1
    return msg_id


async def cdp_send(ws, method, params=None):
    cid = next_id()
    msg = {"id": cid, "method": method}
    if params:
        msg["params"] = params
    await ws.send(json.dumps(msg))
    return cid


async def cdp_recv_until(ws, cid, timeout=15):
    """Receive messages until we get a result for our command id. Return (result, collected_events)."""
    events = []
    deadline = asyncio.get_event_loop().time() + timeout
    while asyncio.get_event_loop().time() < deadline:
        try:
            raw = await asyncio.wait_for(ws.recv(), timeout=2)
            data = json.loads(raw)
            if data.get("id") == cid:
                return data, events
            events.append(data)
        except asyncio.TimeoutError:
            continue
    return None, events


async def drain_events(ws, duration=3):
    """Collect all pending events for a given duration."""
    events = []
    deadline = asyncio.get_event_loop().time() + duration
    while asyncio.get_event_loop().time() < deadline:
        try:
            raw = await asyncio.wait_for(ws.recv(), timeout=0.3)
            data = json.loads(raw)
            events.append(data)
        except asyncio.TimeoutError:
            continue
    return events


def extract_api_from_events(events, requests_map, page_name):
    """Extract API calls from network events into requests_map."""
    for event in events:
        method = event.get("method", "")
        params = event.get("params", {})

        if method == "Network.requestWillBeSent":
            req = params.get("request", {})
            rid = params.get("requestId", "")
            url = req.get("url", "")
            rtype = params.get("type", "")

            # Keep XHR, Fetch, and anything that looks like an API call
            is_api = rtype in ("XHR", "Fetch")
            if not is_api:
                # Also check if URL looks like API
                for pattern in ("/api/", "/agent/", "/member/", "/lottery/", "/fund/",
                                "/bank/", "/bet/", "/game/", "/rebate/", "/user/",
                                "/invite/", "/report/", "/deposit/", "/withdraw"):
                    if pattern in url and not url.endswith((".html", ".js", ".css", ".png", ".jpg", ".gif", ".ico", ".svg", ".woff", ".woff2", ".ttf")):
                        is_api = True
                        break

            if is_api:
                requests_map[rid] = {
                    "page": page_name,
                    "url": url,
                    "method": req.get("method", "GET"),
                    "headers": req.get("headers", {}),
                    "postData": req.get("postData", ""),
                    "type": rtype,
                    "timestamp": str(params.get("timestamp", "")),
                    "response_status": "",
                    "response_headers": {},
                    "response_body": "",
                    "response_mime": "",
                }

        elif method == "Network.responseReceived":
            rid = params.get("requestId", "")
            resp = params.get("response", {})
            if rid in requests_map:
                requests_map[rid]["response_status"] = str(resp.get("status", ""))
                requests_map[rid]["response_headers"] = resp.get("headers", {})
                requests_map[rid]["response_mime"] = resp.get("mimeType", "")


async def get_response_bodies(ws, requests_map):
    """Try to get response bodies for all captured requests."""
    for rid in list(requests_map.keys()):
        try:
            cid = await cdp_send(ws, "Network.getResponseBody", {"requestId": rid})
            result, _ = await cdp_recv_until(ws, cid, timeout=5)
            if result and "result" in result:
                body = result["result"].get("body", "")
                requests_map[rid]["response_body"] = body[:5000]
        except Exception:
            pass


async def main():
    print("=== API Interceptor - Starting ===")

    # Get existing tab with the original site, or use the one we just created
    resp = urllib.request.urlopen(f"{CDP_HTTP}/json/list")
    tabs = json.loads(resp.read().decode())

    # Find original site tab
    target_tab = None
    for tab in tabs:
        if "ee88dly" in tab.get("url", ""):
            target_tab = tab
            break

    if not target_tab:
        print("No tab with original site found. Creating one...")
        req = urllib.request.Request(f"{CDP_HTTP}/json/new?{ORIGINAL_SITE}", method="PUT")
        resp = urllib.request.urlopen(req)
        target_tab = json.loads(resp.read().decode())
        print(f"Created tab: {target_tab['id']}")
    else:
        print(f"Found existing tab: {target_tab['id']} -> {target_tab['url']}")

    ws_url = target_tab["webSocketDebuggerUrl"]
    tab_id = target_tab["id"]

    async with websockets.connect(ws_url, max_size=50 * 1024 * 1024) as ws:
        # Enable domains
        cid = await cdp_send(ws, "Network.enable", {"maxTotalBufferSize": 50_000_000})
        await cdp_recv_until(ws, cid)

        cid = await cdp_send(ws, "Page.enable")
        await cdp_recv_until(ws, cid)

        cid = await cdp_send(ws, "Runtime.enable")
        await cdp_recv_until(ws, cid)

        # Navigate to original site first to make sure we're logged in
        print(f"\nNavigating to {ORIGINAL_SITE}...")
        cid = await cdp_send(ws, "Page.navigate", {"url": ORIGINAL_SITE})
        result, events = await cdp_recv_until(ws, cid)
        await asyncio.sleep(5)
        events += await drain_events(ws, 5)

        # Check if we're on the admin dashboard
        cid = await cdp_send(ws, "Runtime.evaluate", {"expression": "document.title + ' | ' + window.location.href"})
        result, _ = await cdp_recv_until(ws, cid)
        page_info = result.get("result", {}).get("result", {}).get("value", "?") if result else "?"
        print(f"Current page: {page_info}")

        # Capture initial page load APIs
        requests_map = {}
        extract_api_from_events(events, requests_map, "Main Page Load")
        await get_response_bodies(ws, requests_map)
        for rd in requests_map.values():
            all_api_calls.append(rd)
            print(f"  [Main] {rd['method']} {rd['url']}")

        # === Navigate to each page directly ===
        # The original site is iframe-based. Each page is at /agent/{page}.html
        # We'll navigate directly to each page to capture its API calls on load

        for menu in MENU_PAGES:
            page_name = menu["name"]
            page_key = menu["page"]
            page_url = f"{ORIGINAL_SITE}agent/{page_key}.html"

            print(f"\n--- [{page_key}] {page_name} ---")
            print(f"  URL: {page_url}")

            # Clear network cache for clean capture
            cid = await cdp_send(ws, "Network.clearBrowserCache")
            await cdp_recv_until(ws, cid, timeout=3)

            # Navigate to page
            cid = await cdp_send(ws, "Page.navigate", {"url": page_url})
            result, nav_events = await cdp_recv_until(ws, cid)

            # Wait for page to load and API calls to fire
            await asyncio.sleep(4)

            # Collect all events
            all_events = nav_events + await drain_events(ws, 5)
            print(f"  Events collected: {len(all_events)}")

            # Extract API calls
            requests_map = {}
            extract_api_from_events(all_events, requests_map, page_name)

            # Get response bodies
            await get_response_bodies(ws, requests_map)

            for rd in requests_map.values():
                all_api_calls.append(rd)
                print(f"  API: {rd['method']} {rd['url']} -> {rd['response_status']}")

            if not requests_map:
                print(f"  No API calls detected on page load")

            # Now try to trigger search/submit buttons on the page
            # Many pages have a search form that fires an API call
            trigger_js = r"""
            (function() {
                var results = [];
                // Click search buttons
                var btns = document.querySelectorAll('.layui-btn, button[type="submit"], [lay-submit]');
                for (var i = 0; i < btns.length; i++) {
                    var t = btns[i].textContent.trim();
                    if (t.indexOf('\u0054\u00ecm') !== -1 || t.indexOf('search') !== -1 || t.indexOf('Search') !== -1 || t === '' || btns[i].classList.contains('layui-btn-normal')) {
                        btns[i].click();
                        results.push('clicked: ' + t);
                    }
                }
                // Also look for layui form submits
                var forms = document.querySelectorAll('form, .layui-form');
                results.push('forms: ' + forms.length);
                return results.join('; ');
            })()
            """
            cid = await cdp_send(ws, "Runtime.evaluate", {"expression": trigger_js})
            result, trigger_events = await cdp_recv_until(ws, cid)
            trigger_result = result.get("result", {}).get("result", {}).get("value", "") if result else ""
            if trigger_result:
                print(f"  Trigger: {trigger_result}")

            # Collect events from button clicks
            await asyncio.sleep(3)
            click_events = trigger_events + await drain_events(ws, 4)

            click_map = {}
            extract_api_from_events(click_events, click_map, f"{page_name} (search)")

            await get_response_bodies(ws, click_map)

            for rd in click_map.values():
                all_api_calls.append(rd)
                print(f"  API (search): {rd['method']} {rd['url']} -> {rd['response_status']}")

            # Extract API URLs from page JavaScript source code
            extract_js = r"""
            (function() {
                var html = document.documentElement.outerHTML;
                var apis = [];
                // Match $.ajax url patterns
                var re1 = /url\s*:\s*['"]([^'"]+)['"]/g;
                var m;
                while ((m = re1.exec(html)) !== null) {
                    if (m[1].indexOf('/') === 0 || m[1].indexOf('http') === 0) apis.push(m[1]);
                }
                // Match $.get/$.post patterns
                var re2 = /\$\.(get|post)\s*\(\s*['"]([^'"]+)['"]/g;
                while ((m = re2.exec(html)) !== null) {
                    apis.push(m[1].toUpperCase() + ' ' + m[2]);
                }
                // Match fetch patterns
                var re3 = /fetch\s*\(\s*['"]([^'"]+)['"]/g;
                while ((m = re3.exec(html)) !== null) {
                    apis.push(m[1]);
                }
                return JSON.stringify([...new Set(apis)]);
            })()
            """
            cid = await cdp_send(ws, "Runtime.evaluate", {"expression": extract_js})
            result, _ = await cdp_recv_until(ws, cid)
            js_apis_str = result.get("result", {}).get("result", {}).get("value", "[]") if result else "[]"
            try:
                js_apis = json.loads(js_apis_str)
                if js_apis:
                    print(f"  JS-extracted APIs: {len(js_apis)}")
                    for api_url in js_apis:
                        print(f"    {api_url}")
                        # Check if method is embedded
                        parts = api_url.split(" ", 1)
                        if len(parts) == 2 and parts[0] in ("GET", "POST", "PUT", "DELETE"):
                            method_str = parts[0]
                            url_str = parts[1]
                        else:
                            method_str = "GET/POST"
                            url_str = api_url

                        all_api_calls.append({
                            "page": page_name,
                            "url": url_str,
                            "method": method_str,
                            "headers": {},
                            "postData": "",
                            "type": "JS_EXTRACTED",
                            "timestamp": "",
                            "response_status": "N/A (from source)",
                            "response_headers": {},
                            "response_body": "",
                            "response_mime": "",
                        })
            except json.JSONDecodeError:
                pass

    print(f"\n\n=== TOTAL API CALLS CAPTURED: {len(all_api_calls)} ===")
    write_xlsx(all_api_calls)


def write_xlsx(api_calls):
    """Write API documentation to XLSX."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "API Documentation"

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        "STT", "Page/Menu", "API URL", "HTTP Method", "Request Type",
        "Request Headers (key)", "Request Body/Params",
        "Response Status", "Response Content-Type",
        "Response Body (sample)", "Notes"
    ]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    widths = [6, 30, 65, 12, 14, 40, 55, 14, 22, 90, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Deduplicate
    seen = set()
    unique = []
    for c in api_calls:
        key = f"{c['url']}|{c['method']}|{c['page']}"
        if key not in seen:
            seen.add(key)
            unique.append(c)

    data_font = Font(name="Arial", size=10)
    wrap_align = Alignment(vertical="top", wrap_text=True)

    for ri, call in enumerate(unique, 2):
        hdrs_str = ""
        if isinstance(call.get("headers"), dict):
            for k, v in call["headers"].items():
                k_lower = k.lower()
                if k_lower in ("content-type", "authorization", "token", "cookie",
                               "x-requested-with", "accept", "x-token", "x-csrf-token"):
                    hdrs_str += f"{k}: {v}\n"

        post_data = call.get("postData", "")
        if post_data:
            try:
                post_data = json.dumps(json.loads(post_data), indent=2, ensure_ascii=False)
            except (json.JSONDecodeError, TypeError):
                pass

        resp_body = call.get("response_body", "")
        if resp_body:
            try:
                resp_body = json.dumps(json.loads(resp_body), indent=2, ensure_ascii=False)[:3000]
            except (json.JSONDecodeError, TypeError):
                resp_body = resp_body[:3000]

        row_data = [
            ri - 1,
            call.get("page", ""),
            call.get("url", ""),
            call.get("method", ""),
            call.get("type", ""),
            hdrs_str.strip(),
            post_data,
            str(call.get("response_status", "")),
            call.get("response_mime", ""),
            resp_body,
            "",
        ]

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=str(val))
            cell.font = data_font
            cell.alignment = wrap_align
            cell.border = thin_border

    # Summary sheet
    ws2 = wb.create_sheet("API Summary")
    for ci, h in enumerate(["Page", "Count", "API Endpoints"], 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 90

    page_apis = {}
    for c in unique:
        pg = c.get("page", "?")
        page_apis.setdefault(pg, []).append(c.get("url", ""))

    for ri, (pg, urls) in enumerate(page_apis.items(), 2):
        ws2.cell(row=ri, column=1, value=pg)
        ws2.cell(row=ri, column=2, value=len(urls))
        ws2.cell(row=ri, column=3, value="\n".join(urls))
        ws2.cell(row=ri, column=3).alignment = wrap_align

    ws.freeze_panes = "A2"
    ws2.freeze_panes = "A2"
    if len(unique) > 0:
        ws.auto_filter.ref = f"A1:K{len(unique) + 1}"

    output_path = r"c:\Users\Admin\Desktop\MAXHUB\API_Documentation.xlsx"
    wb.save(output_path)
    print(f"\nXLSX saved to: {output_path}")
    print(f"Total unique API calls documented: {len(unique)}")


if __name__ == "__main__":
    asyncio.run(main())
