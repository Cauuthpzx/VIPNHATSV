import paramiko
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# SSH connection
ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
ssh.connect('103.190.80.159', port=43932, username='root', password='MXmgAn#0!l6k')

# Query agent cookies
stdin, stdout, stderr = ssh.exec_command(
    "PGPASSWORD=agenthub123 psql -U agenthub -d agenthub -t -A -F'|' -c "
    "\"SELECT id, name, ext_username, session_cookie, cookie_expires, status, is_active, last_login_at FROM agents ORDER BY id\""
)
err = stderr.read().decode().strip()
if err:
    print("DB Error:", err)
raw = stdout.read().decode().strip()
ssh.close()

if not raw:
    print("No data found")
    exit()

rows = []
for line in raw.split('\n'):
    if line.strip():
        rows.append(line.split('|'))

print(f"Found {len(rows)} agents")

# Create Excel
wb = Workbook()
ws = wb.active
ws.title = "Agent Cookies"

headers = ["ID", "Name", "Username", "Session Cookie", "Cookie Expires", "Status", "Is Active", "Last Login"]
header_fill = PatternFill(start_color="009688", end_color="009688", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

for r, row in enumerate(rows, 2):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')

# Auto-fit column widths
for col in ws.columns:
    max_len = 0
    col_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    # Cap cookie column width
    if col_letter == 'D':  # Session Cookie
        ws.column_dimensions[col_letter].width = 60
    else:
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)

output_path = r"c:\Users\Admin\Desktop\MAXHUB\agent_cookies.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
